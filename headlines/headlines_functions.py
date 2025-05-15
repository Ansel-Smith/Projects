# functions for the headlines project
import openpyxl, os
from datetime import date

os.chdir("/Users/super/Desktop/Projects/webscraping/headlines")

# helper function to check if file has already been uploaded
def finddate(sheet, date):
    for i in range(sheet.max_row):
        if sheet.cell(row=i+1, column=1).value == date:
            return True
    return False

# take all headlines from a txt file and send the data, organized, to an excel spreadsheet
def loader(filename):
    wb = openpyxl.load_workbook('headlines.xlsx')

    date = filename[:filename.find('_')]
    outlet = filename[filename.find('_')+1:filename.find('.')]

    if outlet not in wb.sheetnames:
        wb.create_sheet(outlet)
 
    sheet = wb[outlet]
    
    # break if file has already been uploaded
    if finddate(sheet, date) == True:
        print(filename+" has already been uploaded")
        return

    note = open(outlet+"/"+filename, 'r') 
    rowie=sheet.max_row
    for line in note:
        sheet.cell(row = rowie, column = 1).value = date
        sheet.cell(row = rowie, column = 2).value = line
        sheet.cell(row = rowie, column = 3).value = "=TEXTSPLIT(B+"+str(rowie)+", ' ')"
        rowie = rowie+1
        
    wb.save('headlines.xlsx')
